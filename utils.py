import logging
import os
import math
import datetime
from logging import config
import yaml
import openpyxl
from openpyxl import worksheet
from openpyxl.styles import Font
import sys

log = logging.getLogger(__name__)

def md5(*objs):
    import hashlib
    s = ','.join([str(x) for x in objs])
    hexs = hashlib.md5(s.encode(encoding='utf8')).hexdigest()
    return hexs[:16]

def floor(num:float, ratio):
    ratio = math.pow(10, -ratio)
    return math.floor(num / ratio) * ratio
def ceil(num, ratio):
    ratio = math.pow(10, -ratio)
    return math.ceil(num / ratio) * ratio
def binsearch(l:list, e, key=None):
    if not key:
        key = lambda x:x
    n = len(l)
    bi,ei,mi = 0, n-1, 0
    if e>key(l[ei]) or e<key(l[bi]):
        return -1
    if e == l[ei]:
        return ei
    if e == l[bi]:
        return bi
    while 1:
        mi = math.floor((bi+ei)/2)
        if mi == bi or mi == ei:
            return mi
        if e == key(l[mi]):
            return mi
        elif e > key(l[mi]):
            bi = mi
        else:
            ei = mi

def singleton(cls):
    _instance = {}
    def _singleton(*args, **kargs):
        if cls not in _instance:
            _instance[cls] = cls(*args, **kargs)
        return _instance[cls]
    return _singleton

def ssh_upload(local_path, remote_path, ip, port, user, pwd):
    try:
        import paramiko
        with paramiko.SSHClient() as ssh:
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(ip, port, user, pwd)
            with ssh.open_sftp() as sftp:
                sftp.put(local_path, remote_path)
                log.info('upload file.[%s]->[%s]'%(local_path, '%s@%s:%s'%(user, ip, remote_path)))
        return remote_path
    except BaseException as e:
        log.error('ssh_upload failed.. local[%s] remote[%s]' % (local_path, remote_path), exc_info=e)
        return 
def qwx_msg(url, msg, mentioned_list=None):
    import requests
    data = '''
   {
        "msgtype": "markdown",
        "markdown": {
            "content": "%s"
        }
   }''' % msg.encode("utf-8").decode("latin1")
    headers = {'user-agent': 'my-app/0.0.1'}
    requests.post(url, headers=headers, data=data)

def debug_mode():
    return 'PYDEVD_USE_FRAME_EVAL' in os.environ
__global_pattern = {}
def match_dict(text, pstr=None, pattern = None):
    import re
    if pattern is None:
        if pstr is None: raise BaseException('illegal input.')
        elif pstr in __global_pattern:
            pattern = __global_pattern[pstr]
        else: 
            pattern = re.compile(pstr)
            __global_pattern[pstr] = pattern
    m = pattern.match(text)
    if m is not None:
        return m.groupdict()
################# file ####################
def listFiles(_dir:str = None):
    def recursion(path:str):
        if os.path.isfile(path):
            yield path
        elif os.path.isdir(path):
            for cpath in os.listdir(path):
                for x in recursion(os.path.join(path, cpath)):
                    yield x
    if _dir is None : _dir = os.path.abspath(os.path.dirname(__file__))
    if not os.path.exists(_dir): yield 'done'
    for file in recursion(_dir):
        yield file
def readLines(file:str, encoding='utf8'):
    with open(file, 'r', encoding=encoding, buffering=2<<16) as f:
        while 1:
            line=f.readline()
            if line:
                yield line.strip()
            else:
                return 'done'
def writeLines(file:str, lines, mode='a'):
    with open(file, mode, encoding='utf8', buffering=2<<16) as f:
        for line in lines:
            f.write(str(line) + '\n')
def file_name(path):
    local_path = path.replace('\\', '/')
    fname = local_path[path.rfind('/')+1:]
    return fname
def writeExcel(path:str, rows):
    wb = openpyxl.Workbook()
    wb.guess_types = True
    sheet = wb.active
    for row in rows:
        if isinstance(row, (list, tuple)):
            sheet.append(row)
    defaultStyle(sheet)
    wb.save(path)
def defaultStyle(sheet:worksheet.worksheet.Worksheet):
    column_widths = []
    font=Font(name='微软雅黑')
    for row in sheet:
        for i, cell in enumerate(row):
            celllen = len(str(cell.value))*2+1
            if len(column_widths) > i:
                cell.font = font
                if celllen > column_widths[i]:
                    column_widths[i] = celllen
            else:
                column_widths += [celllen]
    for i, column_width in enumerate(column_widths):
        column_width = min(100, column_width)
        col=sheet.column_dimensions[openpyxl.cell.cell.get_column_letter(i+1)]
        col.width = column_width
def readExcel(path:str, sheet:str=None):
    wb = openpyxl.load_workbook(path)
    table = wb.active if not sheet else wb.get_sheet_by_name(sheet)
    for row in table.rows:
        yield [cell.value for cell in row]
    return row
def read_xlsx(path:str, sheet:str=None):
    import xlrd
    pass
def write_xlsx(path:str, sheet:str=None):
    import xlrd
    pass
##################  http  #######################
def request_regular_fetch(pattern, url, cookies=None, params=None, headers=None):
    try:
        text = request(url, params=params, cookies=cookies, headers = headers)
        import utils
        m = pattern.findall(text)
        if (m is not None) and (len(m) > 0):
            return m
        else: 
            raise ValueError('request_regular_fetch failed. url[%s] pattern[%s]' % (url, pattern))
    except BaseException as e:
        logging.getLogger('dbupdate').error('url fetch failed.url[%s] text:\n%s' % (url, text[:50]), e)
        raise e
def request(url, params=None, cookies=None, headers=None, _type='get'):
    if _type == 'get':
        return Curl().get(url, cookies=cookies, params=params, headers=headers)
    elif _type == 'post':
        return Curl().post(url, cookies=cookies, params=params, headers=headers)
def default_pc_headers():
    return {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9'
        ,'accept-encoding': 'gzip, deflate, br'
        ,'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8'
        ,'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'
    }
def default_phone_headers():
    return {
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9'
        ,'accept-encoding': 'gzip, deflate, br'
        ,'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8'
        ,'sec-ch-ua-platform': "Android"
        ,'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Mobile Safari/537.36'
    }
@singleton
class Cookies:
    def __init__(self):
        _file = 'cookies.yml'
        if not os.path.exists(_file):
            self.datas = {}
        else:
            with open(_file, 'r') as f:
                self.datas = yaml.safe_load(f.read())
        self._file = _file
    def get_all(self, host):
        if host not in self.datas: return {}
        return self.datas[host]
    def get(self, host, key):
        return self.datas[host][key]
    def update(self, host, kws, writeback=True):
        updated = False
        if host not in self.datas:
            self.datas[host] = {}
        for k, v in kws.items():
            if k not in self.datas[host] or self.datas[host] != k:
                updated = True
                self.datas[host][k] = v
        if updated:
            with open(self._file, 'w') as f:
                yaml.dump(self.datas, f)
class Curl:
    def __init__(self):
        pass
    def get(self, url, headers={}, cookies={}, params = {}):
        import requests
        return self.request(url, request_f=requests.get, headers=headers, cookies=cookies, params=params)
    def post(self, url, headers={}, cookies={}, params = {}):
        import requests
        return self.request(url, request_f=requests.post,headers=headers, cookies=cookies, params=params, )
    def request(self, url, request_f, headers={}, cookies={}, params = {}):
        import urllib3
        _, host, _ = urllib3.get_host(url)
        cs = Cookies().get_all(host)
        if cookies is not None: cs |= cookies
        with request_f(url, headers=headers, cookies = cs, params=params) as r:
            if cookies is not None: cs |= r.cookies
            Cookies().update(host, cs)
            bs = r.content
            return str(bs, 'utf8')
##################  timer  #######################
def benchmark(exclude_kw=[], exclude_arg=[], kw_format={}, arg_format={}, threshold=1):
    def str_format(x):
        x_str = str(x).replace('\n',' ')
        if len(x_str) < 150: return x_str
        else: return x_str[:100] + '...' + x_str[-20:]
    def df_format(x):
        return 'df[%d][%d]' % (len(x), len(x.columns))
    def series_format(x):
        return 'series[%d]' % len(x.index)
    def default_str_format(x):
        import pandas as pd
        if isinstance(x, pd.DataFrame): return df_format(x)
        elif isinstance(x, pd.Series): return series_format(x)
        else: return str_format(x)
    def decorator(func):
        def arg_str_format(index, arg):
            if index in exclude_arg: return '_'
            elif index in arg_format: return arg_format[index](arg)
            else: return default_str_format(arg)
        def kw_str_format(k, v):
            if k in exclude_kw: return '%s:_'%k
            elif k in kw_format: return '%s:%s'%(k, kw_format(k))
            else: return '%s:%s'%(k, default_str_format(v))
        def wrapper(*args, **kw):
            def call_str_format():
                str_args = [arg_str_format(id,arg) for id, arg in enumerate(args)]
                str_args += [kw_str_format(k,v) for k, v in kw]
                text = '%s.%s(%s)' % (func.__module__, func.__name__, ', '.join(str_args))
                return text
            b = datetime.datetime.now()
            logging.getLogger('perf').debug('begin: %s'%call_str_format())
            obj = func(*args, **kw)
            e = datetime.datetime.now()
            time_s = (e-b).total_seconds()
            if time_s >= threshold:
                logging.getLogger('perf').info('(%.3fs)call %s' % (time_s, call_str_format()))
            return obj
        return wrapper
    return decorator
import time
class Timer:
    def __init__(self, name:str):
        self._count = 0
        self._elapsed = 0.0
        self._func = time.perf_counter
        self._start = None
        self._name = name
        self._lastcost = 0
    def start(self):
        if self._start is not None: raise RuntimeError('Already started')
        self._start = self._func()
        return self
    def stop(self):
        if self._start is None: raise RuntimeError('Not started')
        end = self._func()
        self._lastcost = end - self._start
        self._elapsed += self._lastcost
        self._count += 1
        self._start = None
        logging.getLogger('perf').info('[%0.3fs %0.3fs/%d=%0.3fs] %s' % (self._lastcost, self._elapsed, self._count, self._elapsed/self._count, self._name))
        return self
    def reset(self):
        self._elapsed = 0.0
        self._count = 0
    @property
    def name(self) -> str : return self._name
    @property
    def running(self) -> bool: return self._start is not None
    @property
    def name(self) -> str: return self._name
    @property
    def count(self) -> int: return self._count
    @property
    def elapsed(self) -> float: return self._elapsed
    @property
    def lastcost(self) -> float: return self._lastcost
    def __enter__(self):
        self.start()
        return self
    def __exit__(self, *args):
        self.stop()
    def __str__(self) -> str:
        pass
__global_timer = {}
def timer(tar):
    if tar in __global_timer: return __global_timer[tar]
    else: 
        t = Timer(tar)
        __global_timer[tar] = t
        return t
def logtime():
    def decorate(func):
        tar = '%s.%s'%(func.__module__, func.__name__)
        def wrapper(*args, **kwargs):
            with timer(tar):
                return func(*args, **kwargs)
        return wrapper
    return decorate

if __name__ == '__main__':
    # ss = listFiles(r'D:\BaiduNetdiskDownload\photos')
    # print(len(list(ss)))
    pass
